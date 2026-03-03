"""
exportar_leads.py — Script de Exportação de Leads QUALIFICADOS

Gera uma planilha Excel (.xlsx) com os leads enriquecidos, prontos para
uso comercial. Permite filtros por classificação e score.
"""

import sys
import argparse
import logging
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Imports internos
from db import get_session, Lead, Processo
from config import OUTPUT_DIR

# ============================================================
# LOGGING
# ============================================================
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("EXPORTADOR")


# ============================================================
# ESTILOS EXCEL
# ============================================================

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER_ALIGNED = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# ============================================================
# EXPORTAÇÃO
# ============================================================

def exportar_xlsx(classificacao: str | None = None, score_min: int = 0, somente_pj: bool = False):
    """
    Consulta os leads no banco SQLite e gera o arquivo XLSX.
    """
    session = get_session()
    
    try:
        # Query com JOIN para pegar dados do Processo
        query = session.query(Lead).join(Processo).filter(Lead.score >= score_min)
        
        if classificacao:
            # Garante que a busca por classificação considere emojis ou texto se houver
            query = query.filter(Lead.classificacao.contains(classificacao.upper()))
            
        if somente_pj:
            query = query.filter(Lead.tipo_pessoa == "Pessoa Jurídica")
            
        leads = query.order_by(Lead.score.desc()).all()
        
        if not leads:
            logger.warning("Nenhum lead encontrado com os filtros aplicados.")
            return None

        # Criar Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads INPI Qualificados"

        # Cabeçalho
        colunas = [
            "Score", "Classificação", "Marca", "Nº Processo", 
            "Titular", "Tipo Pessoa", "UF", "Quantidade de Ataques",
            "Detalhes dos Despachos", "E-mail", "Tipo E-mail", "Telefone", 
            "IPAS", "Dta Depósito", "Classe NICE", "Fonte Enriquecimento"
        ]
        
        for col_num, titulo in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_num, value=titulo)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNED
            cell.border = THIN_BORDER

        # Dados
        for row_num, lead in enumerate(leads, 2):
            # Formatação segura de dados que podem ser None
            marca = lead.processo.marca_nome if lead.processo else "N/A"
            titular = lead.processo.titular_nome if lead.processo else "N/A"
            ipas = lead.processo.codigo_ipas if lead.processo else "N/A"
            deposito = lead.processo.data_deposito.strftime("%d/%m/%Y") if lead.processo and lead.processo.data_deposito else "N/A"
            classe = lead.processo.classe_nice if lead.processo else "N/A"
            
            # Valores das células
            valores = [
                lead.score,
                lead.classificacao,
                marca,
                lead.numero_processo,
                titular,
                lead.tipo_pessoa or "Pessoa Física",
                lead.processo.titular_uf if lead.processo else "N/A",
                lead.quantidade_ataques or 0,
                lead.detalhes_processos or "-",
                lead.email or "-",
                lead.email_tipo or "-",
                lead.telefone or "-",
                ipas,
                deposito,
                classe,
                lead.fonte_enriquecimento or "XML"
            ]
            
            for col_num, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                cell.border = THIN_BORDER
                
                # Cores por classificação (UX pro usuário)
                if col_num == 2: # Coluna Classificação
                    if "QUENTE" in str(valor) or "TIER A" in str(valor):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
                    elif "MORNO" in str(valor) or "TIER B" in str(valor):
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarelo
                
                # Aviso visual para e-mail de escritório
                if col_num == 11 and valor == "escritorio":
                    cell.font = Font(color="FF0000", italic=True)

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)

        # Salvar — sempre sobrescreve o mesmo arquivo (sem acumular)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        save_path = OUTPUT_DIR / "leads_extraidos.xlsx"
        wb.save(save_path)
        
        logger.info(f"✅ Sucesso! {len(leads)} leads exportados para: {save_path}")
        return save_path

    except Exception as e:
        logger.error(f"Erro ao exportar leads: {e}")
        return None
    finally:
        session.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Exporta leads qualificados para XLSX")
    parser.add_argument("--quentes", action="store_true", help="Exportar apenas leads A (Alta Prioridade)")
    parser.add_argument("--score", type=int, default=0, help="Score mínimo para exportação")
    parser.add_argument("--somente-pj", action="store_true", help="Exportar apenas empresas (Pessoa Jurídica)")
    args = parser.parse_args()
    classif = "A (Alta Prioridade)" if args.quentes else None
    
    exportar_xlsx(classificacao=classif, score_min=args.score, somente_pj=args.somente_pj)
